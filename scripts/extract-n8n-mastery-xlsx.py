"""Extract embedded links from the supplied n8n Mastery XLSX catalogue."""

import json
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


INPUT_PATH = Path("/home/ubuntu/upload/n8n-mastery-sources-catalogue.xlsx")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "docs" / "n8n-mastery-xlsx-extracted.json"


def classify_url(url: str | None) -> str:
    if not url:
        return "missing"
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    if "youtube.com" in host or "youtu.be" in host:
        if parsed.path == "/watch" and "v=" in parsed.query or host == "youtu.be":
            return "youtube_direct_video"
        if parsed.path == "/results":
            return "youtube_search"
        return "youtube_other"
    if parsed.scheme in {"http", "https"}:
        return "external_reference"
    return "non_web"


def href(cell) -> str | None:
    if cell.hyperlink is None:
        return None
    return cell.hyperlink.target or cell.hyperlink.location


def main() -> None:
    workbook = load_workbook(INPUT_PATH, read_only=False, data_only=False)
    worksheet = workbook["Master Catalogue"]
    rows: list[dict[str, object]] = []

    for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
        title = values[2].value
        if not title:
            continue
        direct_url = href(values[4])
        rows.append(
            {
                "spreadsheetRow": row_index,
                "catalogueNumber": values[0].value,
                "category": values[1].value,
                "title": title,
                "format": values[3].value,
                "linkLabel": values[4].value,
                "url": direct_url,
                "urlKind": classify_url(direct_url),
                "purpose": values[5].value,
            }
        )

    url_kinds = Counter(str(row["urlKind"]) for row in rows)
    unique_urls = {str(row["url"]) for row in rows if row["url"]}
    result = {
        "input": str(INPUT_PATH),
        "sheet": "Master Catalogue",
        "rowCount": len(rows),
        "uniqueUrlCount": len(unique_urls),
        "urlKinds": dict(sorted(url_kinds.items())),
        "rows": rows,
    }
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({key: result[key] for key in ("rowCount", "uniqueUrlCount", "urlKinds")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
